# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

WEB_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = WEB_ROOT.parent
SERVER_PATH = WEB_ROOT / "server.py"

spec = importlib.util.spec_from_file_location("central_cte_web_server_r124", SERVER_PATH)
server = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(server)


class _TemplateContext:
    def __init__(self, root: Path):
        table_root = root / "uploads" / "tabelas"
        table_root.mkdir(parents=True, exist_ok=True)
        self.upload_categories = {"tabelas": table_root}


class R124QaReportTests(unittest.TestCase):
    def _invoice_pdf(self, path: Path, *, title: str) -> None:
        try:
            from reportlab.pdfgen import canvas
        except ImportError as exc:  # pragma: no cover
            self.skipTest(f"reportlab indisponível: {exc}")
        document = canvas.Canvas(str(path))
        document.setTitle(title)
        lines = (
            "FATURA No: 0000336-0",
            "Transportador: JSP TRANSPORTE E LOGISTICA LTDA",
            "CT-e / RPS / NFS-e",
            "1 001000050465 01/01/26 000286327 1.000,00 10 0,00 76,07",
        )
        y = 800
        for line in lines:
            document.drawString(72, y, line)
            y -= 24
        document.save()

    def test_partner_review_count_ignores_confirmed_ok_statuses(self):
        self.assertFalse(server.partner_review_pending("REVISAR_OK"))
        self.assertFalse(server.partner_review_pending("ALIAS_OK"))
        self.assertFalse(server.partner_review_pending("CONFERIDO"))
        self.assertTrue(server.partner_review_pending("PENDENTE_TRANSCRICAO"))
        self.assertTrue(server.partner_review_pending("REVISÃO MANUAL"))

        scanned = server.scan_partners()
        pending = [rule for rule in scanned["rules"] if rule.get("needs_review")]
        self.assertEqual(len(scanned["rules"]), 679)
        self.assertEqual(len(pending), 0)

    def test_partner_registration_model_is_sanitized_and_guided(self):
        from developer_tools import DeveloperTools
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            self.skipTest(f"openpyxl indisponível: {exc}")

        with tempfile.TemporaryDirectory() as tmp:
            temp = Path(tmp)
            tools = DeveloperTools(PROJECT_ROOT, temp / "data" / "security")
            target = tools.export_partner_registration_template(_TemplateContext(temp))
            workbook = load_workbook(target, read_only=False, data_only=False)
            try:
                self.assertIn("INSTRUCOES", workbook.sheetnames)
                self.assertIn("MODELO_EXEMPLOS", workbook.sheetnames)
                self.assertIn("PARCEIROS", workbook.sheetnames)
                self.assertIn("REGRAS_PERCENTUAL", workbook.sheetnames)
                self.assertEqual(workbook["PARCEIROS"].max_row, 1)
                self.assertEqual(workbook["REGRAS_PERCENTUAL"].max_row, 1)
                self.assertGreater(workbook["MODELO_EXEMPLOS"].max_row, 10)
                self.assertIn("atualização da engine", str(workbook["INSTRUCOES"]["B6"].value).lower())
            finally:
                workbook.close()

    def test_same_engine_document_with_different_pdf_bytes_is_signaled(self):
        from services import OfficialInvoiceEngineService

        with tempfile.TemporaryDirectory() as tmp:
            temp = Path(tmp)
            upload_root = temp / "uploads"
            state_root = temp / "state"
            (upload_root / "bases").mkdir(parents=True)
            first = temp / "A_fatura.pdf"
            second = temp / "B_fatura.pdf"
            self._invoice_pdf(first, title="Original")
            self._invoice_pdf(second, title="Cópia com metadados diferentes")
            self.assertNotEqual(first.read_bytes(), second.read_bytes())

            service = OfficialInvoiceEngineService(PROJECT_ROOT, upload_root, state_root)
            summary = service.process([first, second])
            records = {item["file"]: item for item in service.stored_file_records([first, second])}

        self.assertEqual(summary["processed_files"], 1)
        self.assertEqual(summary["duplicate_files"], 1)
        duplicate = records[second.name]
        self.assertEqual(duplicate["status"], "duplicate")
        self.assertEqual(duplicate["code"], "DUPLICATE_ENGINE_DOCUMENT")
        self.assertEqual(duplicate["duplicate_of"], first.name)
        self.assertIn("rejeitada por duplicidade", duplicate["reason"].lower())
        self.assertIn("excluído do cálculo", duplicate["financial_disposition"].lower())

    def test_signature_filter_and_pdf_fallback_remain_corrected(self):
        css = (WEB_ROOT / "static" / "styles.css").read_text(encoding="utf-8")
        dockerfile = (PROJECT_ROOT / "deploy" / "vps" / "Dockerfile").read_text(encoding="utf-8")
        signature_service = (WEB_ROOT / "services" / "official_signature_service.py").read_text(encoding="utf-8")
        self.assertIn("grid-template-columns:minmax(0,1fr) minmax(112px,128px)", css)
        self.assertIn(".signature-filterbar select { min-width:0; width:100%; max-width:128px; }", css)
        self.assertIn("poppler-utils", dockerfile)
        self.assertIn("pdftoppm", signature_service)


if __name__ == "__main__":
    unittest.main()
