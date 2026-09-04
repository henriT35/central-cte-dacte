# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib.util
import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

WEB_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = WEB_ROOT.parent
SERVER_PATH = WEB_ROOT / "server.py"

spec = importlib.util.spec_from_file_location("central_cte_web_server_mvp13", SERVER_PATH)
server = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(server)

from developer_tools import DeveloperTools  # noqa: E402
from security import AuthManager, AuthenticatedUser  # noqa: E402


class FakeXmlService:
    def __init__(self, source: Path):
        self.source = Path(source)
        self.invalidated = 0
        self.cleared = 0

    def resolve_table_source(self, *, raise_on_missing=True):
        if self.source.is_file():
            return self.source
        if raise_on_missing:
            raise FileNotFoundError("Tabela ausente")
        return None

    def validate_table_source(self, source: Path):
        workbook = load_workbook(source, read_only=False, data_only=False)
        try:
            partners = workbook["PARCEIROS"]
            headers = [str(cell.value or "").strip() for cell in partners[1]]
            partner_col = headers.index("Parceiro ID") + 1
            partner_ids = {
                str(partners.cell(row, partner_col).value or "").strip()
                for row in range(2, partners.max_row + 1)
                if str(partners.cell(row, partner_col).value or "").strip()
            }
            rules = workbook["REGRAS_PERCENTUAL"]
            rule_count = sum(
                1
                for row in range(2, rules.max_row + 1)
                if any(rules.cell(row, col).value not in (None, "") for col in range(1, rules.max_column + 1))
            )
            return {
                "partner_count": len(partner_ids),
                "rule_count": rule_count,
                "source": str(source),
            }
        finally:
            workbook.close()

    def invalidate_dependencies(self):
        self.invalidated += 1

    def clear_results(self):
        self.cleared += 1
        return {"cleared": True}


class Mvp13GovernanceTests(unittest.TestCase):
    def test_all_authenticated_users_can_report_but_only_developer_reads_global_notebook(self):
        with tempfile.TemporaryDirectory() as tmp:
            tools = DeveloperTools(PROJECT_ROOT, Path(tmp) / "security")
            admin = AuthenticatedUser("a", "admin", "Admin", "admin")
            operator = AuthenticatedUser("o", "op", "Operador", "operador")
            reader = AuthenticatedUser("r", "consulta", "Consulta", "consulta")
            developer = AuthenticatedUser("d", "dev", "Dev", "desenvolvedor")
            for user in (admin, operator, reader, developer):
                self.assertTrue(tools.capabilities(user)["can_submit_qa"])
            for user in (admin, operator, reader):
                self.assertFalse(tools.capabilities(user)["can_view_qa"])
                self.assertFalse(tools.capabilities(user)["can_view_infrastructure"])
            self.assertTrue(tools.capabilities(developer)["can_view_qa"])
            self.assertTrue(tools.capabilities(developer)["can_clear_qa"])
            self.assertTrue(tools.capabilities(developer)["can_view_infrastructure"])

    def test_qa_migration_consolidates_workspaces_in_one_global_notebook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workspaces = root / "workspaces"
            global_root = root / "qa"
            first = workspaces / "user-a" / "state" / "qa_notes.json"
            second = workspaces / "user-b" / "state" / "qa_notes.json"
            first.parent.mkdir(parents=True)
            second.parent.mkdir(parents=True)
            first.write_text(json.dumps([
                {"id": "QA-1", "title": "A", "updated_at": "2026-07-27T10:00:00-03:00"},
            ]), encoding="utf-8")
            second.write_text(json.dumps([
                {"id": "QA-1", "title": "A atualizado", "updated_at": "2026-07-27T11:00:00-03:00"},
                {"id": "QA-2", "title": "B", "updated_at": "2026-07-27T09:00:00-03:00"},
            ]), encoding="utf-8")

            original = (
                server.WORKSPACES_ROOT,
                server.GLOBAL_QA_ROOT,
                server.GLOBAL_QA_PATH,
                server.QA_MIGRATION_DONE,
            )
            try:
                server.WORKSPACES_ROOT = workspaces
                server.GLOBAL_QA_ROOT = global_root
                server.GLOBAL_QA_PATH = global_root / "qa_notes.json"
                server.QA_MIGRATION_DONE = False
                result = server.migrate_workspace_qa_notes()
                merged = json.loads(server.GLOBAL_QA_PATH.read_text(encoding="utf-8"))
            finally:
                (
                    server.WORKSPACES_ROOT,
                    server.GLOBAL_QA_ROOT,
                    server.GLOBAL_QA_PATH,
                    server.QA_MIGRATION_DONE,
                ) = original

            self.assertEqual(result["total"], 2)
            self.assertEqual({row["id"] for row in merged}, {"QA-1", "QA-2"})
            self.assertEqual(next(row for row in merged if row["id"] == "QA-1")["title"], "A atualizado")
            self.assertTrue(first.with_name("qa_notes_migrated_mvp13.json").is_file())
            self.assertTrue(second.with_name("qa_notes_migrated_mvp13.json").is_file())

    def test_only_developer_can_edit_or_delete_users_with_safety_guards(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            auth = AuthManager(root)
            developer = auth.setup_developer("dev.local", "Desenvolvedor", "SenhaSegura123")
            admin = auth.create_user("admin.local", "Administrador", "admin", "OutraSenha123", actor=developer)
            operator = auth.create_user("operador.local", "Operador", "operador", "TerceiraSenha123", actor=admin)

            with self.assertRaises(PermissionError):
                auth.update_user(operator.id, display_name="Novo", role="consulta", active=True, actor=admin)
            updated = auth.update_user(
                operator.id,
                display_name="Operador Editado",
                role="consulta",
                active=False,
                actor=developer,
            )
            self.assertEqual(updated.display_name, "Operador Editado")
            self.assertEqual(updated.role, "consulta")

            with self.assertRaises(PermissionError):
                auth.update_user(developer.id, display_name="Dev", role="admin", active=True, actor=developer)
            with self.assertRaises(PermissionError):
                auth.delete_user(developer.id, actor=developer)

            result = auth.delete_user(operator.id, actor=developer)
            self.assertTrue(result["deleted"])
            self.assertTrue(result["workspace_preserved"])
            self.assertTrue((root / "deleted_users.json").is_file())
            self.assertNotIn(operator.id, {row["id"] for row in auth.list_users(actor=developer)})

    def test_partner_table_is_split_into_one_workbook_per_partner_and_can_be_deleted(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            table_upload = root / "workspace" / "uploads" / "tabelas"
            table_upload.mkdir(parents=True)
            legacy = table_upload / "cadastro_tabelas_parceiros.xlsx"

            from openpyxl import Workbook
            workbook = Workbook()
            partners = workbook.active
            partners.title = "PARCEIROS"
            partners.append(["Parceiro ID", "Nome Parceiro", "Status"])
            partners.append(["P1", "Parceiro Um", "ATIVO"])
            partners.append(["P2", "Parceiro Dois", "ATIVO"])
            partners.append(["P3", "Parceiro Três", "ATIVO"])
            rules = workbook.create_sheet("REGRAS_PERCENTUAL")
            rules.append(["Regra ID", "Parceiro ID", "Destino Cidade", "Percentual"])
            rules.append(["R1", "P1", "MANAUS", 10])
            rules.append(["R2", "P2", "BELEM", 20])
            rules.append(["R3", "P3", "SANTAREM", 30])
            config = workbook.create_sheet("CONFIG_PROGRAMA")
            config.append(["Chave", "Valor", "Descrição"])
            config.append(["VERSAO", "1", "Teste"] )
            workbook.save(legacy)
            workbook.close()

            xml = FakeXmlService(legacy)
            context = SimpleNamespace(
                upload_categories={"tabelas": table_upload, "bases": root / "workspace" / "uploads" / "bases"},
                state_root=root / "workspace" / "state",
                xml_service=xml,
                qa_path=root / "qa" / "qa_notes.json",
            )
            tools = DeveloperTools(PROJECT_ROOT, root / "data" / "security")

            overview = tools.ensure_partner_files(context)
            rows = tools.partner_files_overview(context)
            self.assertEqual(overview["file_count"], 3)
            self.assertEqual(len(rows), 3)
            self.assertTrue(tools.partner_aggregate_path.is_file())

            for row in rows:
                partner_book = load_workbook(row["path"], read_only=False, data_only=False)
                try:
                    sheet = partner_book["PARCEIROS"]
                    headers = [str(cell.value or "").strip() for cell in sheet[1]]
                    partner_col = headers.index("Parceiro ID") + 1
                    ids = {
                        str(sheet.cell(index, partner_col).value or "").strip()
                        for index in range(2, sheet.max_row + 1)
                        if str(sheet.cell(index, partner_col).value or "").strip()
                    }
                    self.assertEqual(ids, {row["partner_id"]})
                finally:
                    partner_book.close()

            export_path = tools.export_partner_files_zip(context)
            with zipfile.ZipFile(export_path, "r") as archive:
                self.assertEqual(len([name for name in archive.namelist() if name.endswith(".xlsx")]), 3)

            deleted = tools.delete_partner_file(context, rows[-1]["partner_id"])
            self.assertTrue(deleted["deleted"])
            self.assertEqual(len(tools.partner_files_overview(context)), 2)
            self.assertTrue(Path(deleted["backup"]).is_file())

    def test_interface_hides_infrastructure_and_management_actions_from_nondevelopers(self):
        page = (WEB_ROOT / "static" / "index.html").read_text(encoding="utf-8")
        script = (WEB_ROOT / "static" / "app.js").read_text(encoding="utf-8")
        self.assertIn('id="environment-local-card" class="panel settings-card developer-only hidden"', page)
        self.assertIn('id="server-domain-card" class="panel settings-card operations-card developer-only hidden"', page)
        self.assertIn('id="user-edit-modal"', page)
        self.assertIn('id="partner-file-list"', page)
        self.assertIn("qa-submit-access", page)
        self.assertIn("qa-view-access", page)
        self.assertIn('$$(".developer-only")', script)
        self.assertIn("can_view_infrastructure", (WEB_ROOT / "developer_tools.py").read_text(encoding="utf-8"))
        self.assertIn("openUserEditModal", script)
        self.assertIn("deletePartnerFile", script)


if __name__ == "__main__":
    unittest.main()
