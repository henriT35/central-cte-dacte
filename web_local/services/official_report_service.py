# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib
import json
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable, Mapping

from .xml_report_web_patch import (
    correct_report_model,
    patch_report_xlsx,
    prepare_report_files,
)

SERVICE_VERSION = "2.7.0 RC27.14 WEB/WINDOWS MVP13 R12.7"


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def read_json(path: Path, fallback: Any) -> Any:
    try:
        if path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return fallback


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    temporary.replace(path)


def _paths(paths: Iterable[Path], suffix: str) -> list[Path]:
    return sorted(
        {
            Path(path).resolve()
            for path in paths
            if Path(path).is_file() and Path(path).suffix.lower() == suffix
        },
        key=lambda path: str(path).lower(),
    )


def _money_equal(left: Any, right: Any, tolerance: float = 0.01) -> bool:
    try:
        return abs(float(left or 0.0) - float(right or 0.0)) <= tolerance
    except Exception:
        return False


def _append_invoice_file_alerts_sheet(target: Path, file_records: Iterable[Mapping[str, Any]]) -> int:
    alerts = [
        dict(item)
        for item in file_records
        if str(item.get("status") or "") in {"rejected", "duplicate", "received", "processing"}
    ]
    if not alerts:
        return 0

    vendor_root = Path(__file__).resolve().parents[1] / "vendor"
    if vendor_root.is_dir() and str(vendor_root) not in sys.path:
        sys.path.insert(0, str(vendor_root))
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if target.is_file():
        workbook = load_workbook(target)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
    title = "ARQUIVOS_REJEITADOS"
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title)
    headers = [
        "Arquivo",
        "Status",
        "Etapa",
        "Código",
        "Motivo",
        "Duplicado de",
        "Base da duplicidade",
        "Tratamento financeiro",
        "Fatura(s)",
        "Parceiro(s)",
        "Leitor PDF",
        "SHA-256",
        "Tamanho (bytes)",
        "Recebido em",
        "Processado em",
    ]
    sheet.append(headers)
    status_labels = {
        "rejected": "REJEITADO",
        "duplicate": "DUPLICADO",
        "received": "NÃO PROCESSADO",
        "processing": "PROCESSAMENTO INCOMPLETO",
    }
    for item in alerts:
        sheet.append([
            str(item.get("file") or ""),
            status_labels.get(str(item.get("status") or ""), str(item.get("status") or "").upper()),
            str(item.get("stage") or ""),
            str(item.get("code") or ""),
            str(item.get("reason") or ""),
            str(item.get("duplicate_of") or ""),
            str(item.get("duplicate_basis") or ""),
            str(item.get("financial_disposition") or ""),
            ", ".join(str(value) for value in (item.get("invoice_numbers") or item.get("invoice_keys") or [])),
            ", ".join(str(value) for value in (item.get("partners") or [])),
            str(item.get("text_backend") or ""),
            str(item.get("sha256") or ""),
            int(item.get("size_bytes") or 0),
            str(item.get("received_at") or ""),
            str(item.get("processed_at") or ""),
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[7].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            if cell.column not in {5, 8}:
                cell.alignment = Alignment(vertical="top")
    widths = {
        "A": 34,
        "B": 30,
        "C": 20,
        "D": 30,
        "E": 70,
        "F": 34,
        "G": 26,
        "H": 52,
        "I": 22,
        "J": 28,
        "K": 18,
        "L": 68,
        "M": 18,
        "N": 24,
        "O": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    workbook.save(target)
    return len(alerts)


class OfficialReportService:
    """Gera os XLSX oficiais a partir dos snapshots já publicados pelo motor.

    Esta camada não recalcula frete, pedágio, GRIS ou decisões financeiras.
    Ela apenas reconstrói a entrada esperada pelos consolidadores oficiais
    RC26.6 e chama os escritores OpenXML já homologados pelo projeto.
    """

    def __init__(
        self,
        project_root: Path,
        upload_root: Path,
        output_root: Path,
        state_root: Path,
        xml_service: Any,
        invoice_service: Any,
    ) -> None:
        self.project_root = Path(project_root).resolve()
        self.upload_root = Path(upload_root).resolve()
        self.output_root = Path(output_root).resolve() / "relatorios"
        self.state_root = Path(state_root).resolve()
        self.xml_service = xml_service
        self.invoice_service = invoice_service
        self.last_run_path = self.state_root / "report_generation_last_run.json"
        self._lock = threading.RLock()
        self._xml_report_module: Any | None = None

    @property
    def engine_root(self) -> Path:
        return self.project_root / "engine"

    @property
    def xml_report_file(self) -> Path:
        return self.engine_root / "central_cte_modular" / "reports" / "xml_validation_report.py"

    @property
    def invoice_report_file(self) -> Path:
        return self.engine_root / "central_cte_modular" / "reports" / "invoice_report.py"

    def readiness(self) -> dict[str, Any]:
        ready = self.xml_report_file.is_file() and self.invoice_report_file.is_file()
        return {
            "connected": bool(ready),
            "service_version": SERVICE_VERSION,
            "status": (
                "Geradores XLSX oficiais disponíveis sem dependência da interface antiga."
                if ready
                else "Geradores oficiais de relatório não foram encontrados."
            ),
            "output_root": str(self.output_root),
            "last_run": read_json(self.last_run_path, {}),
        }

    def _load_xml_report_module(self) -> Any:
        with self._lock:
            if self._xml_report_module is not None:
                return self._xml_report_module
            engine_path = str(self.engine_root)
            if engine_path not in sys.path:
                sys.path.insert(0, engine_path)
            self._xml_report_module = importlib.import_module(
                "central_cte_modular.reports.xml_validation_report"
            )
            return self._xml_report_module

    def _target(self, prefix: str) -> Path:
        self.output_root.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = self.output_root / f"{prefix}_{stamp}.xlsx"
        counter = 2
        while target.exists():
            target = self.output_root / f"{prefix}_{stamp}_{counter}.xlsx"
            counter += 1
        return target

    def generate_xml(self, xml_paths: Iterable[Path]) -> dict[str, Any]:
        started = time.monotonic()
        paths = _paths(xml_paths, ".xml")
        if not paths:
            raise ValueError("Adicione e processe pelo menos um XML antes de gerar o relatório.")

        stale: list[str] = []
        report_files: list[dict[str, Any]] = []
        for path in paths:
            stored = self.xml_service.stored_row(path)
            if not isinstance(stored, Mapping):
                stale.append(path.name)
                continue
            info = dict(stored.get("engine_info") or {})
            validation = dict(stored.get("validation") or {})
            if not info or not validation:
                stale.append(path.name)
                continue
            manual = stored.get("manual_decision")
            if isinstance(manual, Mapping) and manual.get("decision"):
                decision = str(manual.get("decision") or "").strip().lower()
                labels = {"approved": "APROVADO", "rejected": "RECUSADO", "pending": "PENDENTE"}
                info["revisao_manual"] = labels.get(decision, decision.upper())
                info["observacao_manual"] = str(manual.get("reason") or stored.get("manual_reason") or "")
                info["revisao_data"] = str(manual.get("decided_at") or stored.get("manual_decided_at") or "")
                info["revisao_usuario"] = str(manual.get("actor_name") or manual.get("actor_id") or "")
            info["validacao"] = validation
            info["path"] = str(path)
            info["arquivo"] = str(path)
            report_files.append(info)

        if stale:
            preview = ", ".join(stale[:4])
            suffix = "…" if len(stale) > 4 else ""
            raise ValueError(
                "Há XMLs sem fotografia oficial válida. Processe novamente a validação antes do relatório: "
                f"{preview}{suffix}"
            )

        engine = self.xml_service._load_engine()
        export_row = getattr(engine, "validation_export_row", None)
        module = self._load_xml_report_module()
        generator_type = getattr(module, "XmlValidationReportGenerator", None)
        if generator_type is None:
            raise RuntimeError("O consolidator oficial do relatório XML não foi publicado.")
        generator = generator_type(export_row if callable(export_row) else None)
        target = self._target("relatorio_validacao_xml_RC26_6_WEB")
        report_files = prepare_report_files(report_files)
        model = generator.build(report_files)
        correct_report_model(module, model)
        generator.writer.write(target, model)
        patch_report_xlsx(target, module, model)
        generator.last_model = model
        if not target.is_file() or target.stat().st_size <= 0:
            raise RuntimeError("O escritor oficial não produziu o arquivo XLSX de XMLs.")
        if int(getattr(model, "source_count", 0) or 0) != len(report_files):
            raise RuntimeError("O relatório XML não preservou a quantidade de documentos processados.")

        metrics = dict(getattr(model, "metrics", {}) or {})
        result = {
            "status": "concluido",
            "module": "xml",
            "generated_at": now_iso(),
            "elapsed_seconds": round(time.monotonic() - started, 3),
            "path": str(target),
            "name": target.name,
            "size_bytes": target.stat().st_size,
            "documents": len(report_files),
            "attention_rows": len(getattr(model, "attention_rows", []) or []),
            "detail_rows": len(getattr(model, "detail_rows", []) or []),
            "audit_rows": len(getattr(model, "audit_rows", []) or []),
            "metrics": metrics,
            "writer": "XmlValidationReportGenerator RC26.6 + publicação web MVP13 R1",
        }
        write_json_atomic(self.last_run_path, result)
        return result

    def generate_invoices(
        self,
        pdf_paths: Iterable[Path],
        *,
        only_problem_invoices: bool = False,
    ) -> dict[str, Any]:
        started = time.monotonic()
        paths = _paths(pdf_paths, ".pdf")
        if not paths:
            raise ValueError("Adicione e processe pelo menos uma fatura PDF antes de gerar o relatório.")

        stored_rows = self.invoice_service.stored_rows(paths)
        payload = self.invoice_service._stored_payload_for_paths(paths)
        if not payload:
            raise ValueError("As faturas mudaram ou ainda não foram processadas. Processe-as novamente.")
        file_records = payload.get("file_records") if isinstance(payload, Mapping) else []
        if not stored_rows:
            target = self._target("relatorio_faturas_rejeitadas_RC26_6_WEB")
            rejected_sheet_rows = _append_invoice_file_alerts_sheet(
                target,
                file_records if isinstance(file_records, list) else [],
            )
            if not rejected_sheet_rows:
                raise ValueError("Nenhuma fatura válida ou arquivo rejeitado foi encontrado para o relatório.")
            stored_summary = payload.get("summary") if isinstance(payload, Mapping) else {}
            if not isinstance(stored_summary, Mapping):
                stored_summary = {}
            result = {
                "status": "concluido_com_alertas",
                "module": "invoices",
                "only_problem_invoices": bool(only_problem_invoices),
                "generated_at": now_iso(),
                "elapsed_seconds": round(time.monotonic() - started, 3),
                "path": str(target),
                "name": target.name,
                "size_bytes": target.stat().st_size,
                "invoices": 0,
                "items": 0,
                "problem_items": 0,
                "total_value": 0.0,
                "payable_value": 0.0,
                "future_value": 0.0,
                "internal_problem_value": 0.0,
                "rejected_file_rows": rejected_sheet_rows,
                "rejected_file_sheet": "ARQUIVOS_REJEITADOS",
                "writer": "Exportação operacional R12 de arquivos rejeitados; motor RC26.6 sem decisão financeira",
            }
            write_json_atomic(self.last_run_path, result)
            return result
        records = payload.get("report_records") if isinstance(payload, Mapping) else None
        if not isinstance(records, list) or not records:
            raise ValueError(
                "A fotografia técnica do relatório não existe nesta execução. "
                "Processe novamente as faturas nesta versão antes de gerar o XLSX."
            )
        report_records = [dict(record) for record in records if isinstance(record, Mapping)]
        if not report_records:
            raise ValueError("Nenhum registro financeiro oficial foi encontrado para o relatório.")

        page = SimpleNamespace(
            invoice_detail_records=report_records,
            invoice_rows=[
                [row.get("invoice") or row.get("invoice_key") or "", row.get("partner") or ""]
                for row in stored_rows
            ],
            _last_decision_snapshot=None,
        )
        services, _build_read_model = self.invoice_service._load_services()
        build_result = services.report_builder.build(page, bool(only_problem_invoices))

        stored_summary = payload.get("summary") if isinstance(payload, Mapping) else {}
        if not isinstance(stored_summary, Mapping):
            stored_summary = {}
        contract = {
            "total_value": (getattr(build_result, "total_value", 0.0), stored_summary.get("total_value")),
            "payable_value": (getattr(build_result, "payable_value", 0.0), stored_summary.get("payable_value")),
            "future_value": (getattr(build_result, "future_value", 0.0), stored_summary.get("future_value")),
            "blocked_value": (getattr(build_result, "blocked_value", 0.0), stored_summary.get("internal_problem_value")),
        }
        mismatches = [
            name for name, (observed, expected) in contract.items()
            if not _money_equal(observed, expected)
        ]
        if mismatches and not only_problem_invoices:
            raise RuntimeError(
                "O relatório de faturas divergiu da fotografia financeira oficial nos campos: "
                + ", ".join(mismatches)
            )

        prefix = (
            "relatorio_faturas_problemas_RC26_6_WEB"
            if only_problem_invoices
            else "relatorio_faturas_RC26_6_WEB"
        )
        target = self._target(prefix)
        services.xlsx_writer.write(target, build_result.sheets)
        if not target.is_file() or target.stat().st_size <= 0:
            raise RuntimeError("O escritor oficial não produziu o arquivo XLSX de faturas.")

        rejected_sheet_rows = _append_invoice_file_alerts_sheet(
            target,
            file_records if isinstance(file_records, list) else [],
        )

        log_path = getattr(services.xlsx_writer, "last_log_path", None)
        result = {
            "status": "concluido_com_alertas" if rejected_sheet_rows else "concluido",
            "module": "invoices",
            "only_problem_invoices": bool(only_problem_invoices),
            "generated_at": now_iso(),
            "elapsed_seconds": round(time.monotonic() - started, 3),
            "path": str(target),
            "name": target.name,
            "size_bytes": target.stat().st_size,
            "invoices": int(getattr(build_result, "invoice_count", 0) or 0),
            "items": int(getattr(build_result, "item_count", 0) or 0),
            "problem_items": int(getattr(build_result, "problem_item_count", 0) or 0),
            "total_value": float(getattr(build_result, "total_value", 0.0) or 0.0),
            "payable_value": float(getattr(build_result, "payable_value", 0.0) or 0.0),
            "future_value": float(getattr(build_result, "future_value", 0.0) or 0.0),
            "internal_problem_value": float(getattr(build_result, "blocked_value", 0.0) or 0.0),
            "decision_log": str(log_path) if log_path else "",
            "rejected_file_rows": rejected_sheet_rows,
            "rejected_file_sheet": "ARQUIVOS_REJEITADOS" if rejected_sheet_rows else "",
            "writer": "InvoiceReportBuilder / InvoiceExecutiveXlsxWriter RC26.6",
        }
        write_json_atomic(self.last_run_path, result)
        return result

    def generate(
        self,
        module: str,
        *,
        xml_paths: Iterable[Path] = (),
        pdf_paths: Iterable[Path] = (),
        only_problem_invoices: bool = False,
    ) -> dict[str, Any]:
        normalized = str(module or "").strip().lower()
        with self._lock:
            if normalized == "xml":
                return self.generate_xml(xml_paths)
            if normalized in {"invoices", "faturas"}:
                return self.generate_invoices(
                    pdf_paths,
                    only_problem_invoices=only_problem_invoices,
                )
            raise ValueError("Módulo de relatório inválido. Use xml ou invoices.")


__all__ = ["OfficialReportService", "SERVICE_VERSION"]
