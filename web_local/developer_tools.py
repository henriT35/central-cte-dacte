# -*- coding: utf-8 -*-
from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import time
import uuid
import zipfile
import threading
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping
from xml.etree import ElementTree as ET

DEFAULT_FEATURE_FLAGS: dict[str, bool] = {
    "qa_admin": False,
    "qa_operador": False,
    "qa_consulta": False,
    "beta_admin": False,
    "beta_operador": False,
    "technical_reports_admin": False,
}

BATCH_PATTERN = re.compile(r"^[A-Za-z0-9_-]{8,80}$")
MAX_BASE_FILE_BYTES = 220 * 1024 * 1024
MAX_BASE_FILES = 60
MAX_PARTNER_TABLE_BYTES = 40 * 1024 * 1024


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def safe_name(value: Any, fallback: str) -> str:
    name = Path(str(value or fallback).replace("\\", "/")).name.strip().replace("\x00", "")
    name = re.sub(r"[^A-Za-z0-9À-ÿ._()\- \[\]]+", "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return (name or fallback)[:180]


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_json(path: Path, fallback: Any) -> Any:
    try:
        if path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return fallback


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    temporary.replace(path)


class DeveloperTools:
    def __init__(self, project_root: Path, security_root: Path):
        self.project_root = Path(project_root).resolve()
        self.security_root = Path(security_root).resolve()
        self.features_path = self.security_root / "developer_features.json"
        self.global_data_root = self.security_root.parent
        self.partner_root = self.global_data_root / "partner_tables"
        self.partner_files_root = self.partner_root / "files"
        self.partner_history_root = self.partner_root / "history"
        self.partner_export_root = self.partner_root / "exports"
        self.partner_aggregate_path = self.partner_root / "cadastro_tabelas_parceiros_compilada.xlsx"
        self.qa_root = self.global_data_root / "qa"
        self.qa_attachment_root = self.qa_root / "attachments"
        self.qa_export_root = self.qa_root / "exports"
        self.workspaces_root = self.global_data_root / "workspaces"
        self._partner_lock = threading.RLock()
        for directory in (
            self.partner_root,
            self.partner_files_root,
            self.partner_history_root,
            self.partner_export_root,
            self.qa_root,
            self.qa_attachment_root,
            self.qa_export_root,
        ):
            directory.mkdir(parents=True, exist_ok=True)

    def load_features(self) -> dict[str, bool]:
        raw = read_json(self.features_path, {})
        result = dict(DEFAULT_FEATURE_FLAGS)
        if isinstance(raw, Mapping):
            for key in result:
                if key in raw:
                    result[key] = bool(raw[key])
        return result

    def save_features(self, payload: Mapping[str, Any]) -> dict[str, bool]:
        current = self.load_features()
        for key in current:
            if key in payload:
                current[key] = bool(payload[key])
        write_json_atomic(self.features_path, current)
        return current

    def capabilities(self, user: Any | None) -> dict[str, bool]:
        role = str(getattr(user, "role", "") or "consulta").strip().lower()
        flags = self.load_features()
        developer = role == "desenvolvedor"
        admin = role == "admin"
        return {
            "is_developer": developer,
            "can_manage_users": developer,
            "can_edit_users": developer,
            "can_create_developer": developer,
            "can_import_base": role in {"admin", "desenvolvedor"},
            "can_override_xml_status": role in {"admin", "desenvolvedor"},
            "can_manage_partner_tables": developer,
            "can_manage_features": developer,
            "can_submit_qa": bool(user),
            "can_view_qa": developer,
            "can_clear_qa": developer,
            "can_view_infrastructure": developer,
            "can_view_security_readiness": developer,
            "can_view_technical_reports": developer,
            "beta_enabled": developer or bool(flags.get(f"beta_{role}", False)),
            "can_manage_backups": developer,
            "can_manage_database_integration": developer,
            "can_view_audit": role in {"admin", "desenvolvedor"},
        }

    @staticmethod
    def _file_rows(paths: list[Path]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for path in sorted(paths, key=lambda item: item.name.lower()):
            if not path.is_file():
                continue
            stat = path.stat()
            rows.append({
                "name": path.name,
                "path": str(path.resolve()),
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
                "sha256": sha256_file(path),
            })
        return rows

    def base_overview(self, context: Any) -> dict[str, Any]:
        uploaded_root = Path(context.upload_categories["bases"]).resolve()
        project_root = self.project_root / "bases"
        uploaded_files = sorted(uploaded_root.glob("*.sswweb")) if uploaded_root.is_dir() else []
        active_root = uploaded_root if uploaded_files else project_root
        active_files = sorted(active_root.glob("*.sswweb")) if active_root.is_dir() else []
        return {
            "source": "importada" if uploaded_files else "embutida",
            "active_root": str(active_root),
            "active_files": self._file_rows(active_files),
            "active_file_count": len(active_files),
            "uploaded_file_count": len(uploaded_files),
            "replace_mode": "conjunto completo",
            "message": "Ao confirmar uma nova base, todo o conjunto importado anterior é substituído de forma atômica.",
        }

    @staticmethod
    def _validate_batch_id(value: Any) -> str:
        batch_id = str(value or "").strip()
        if not BATCH_PATTERN.fullmatch(batch_id):
            raise ValueError("Identificador de lote da Base SSW inválido.")
        return batch_id

    def stage_base_file(self, context: Any, batch_id: Any, filename: Any, payload: bytes) -> dict[str, Any]:
        batch = self._validate_batch_id(batch_id)
        name = safe_name(filename, "base.sswweb")
        if Path(name).suffix.lower() != ".sswweb":
            raise ValueError("A Base SSW aceita somente arquivos .sswweb.")
        if not payload or len(payload) > MAX_BASE_FILE_BYTES:
            raise ValueError("Cada arquivo da Base SSW deve ter conteúdo e no máximo 220 MB.")
        sample = payload[:16384]
        if b"\x00" in sample or b";" not in sample:
            raise ValueError("O arquivo não parece ser uma planilha SSW Web válida.")
        stage = Path(context.state_root) / "base_staging" / batch
        stage.mkdir(parents=True, exist_ok=True)
        existing = list(stage.glob("*.sswweb"))
        if len(existing) >= MAX_BASE_FILES and not (stage / name).exists():
            raise ValueError(f"O lote ultrapassa o limite de {MAX_BASE_FILES} arquivos.")
        destination = stage / name
        destination.write_bytes(payload)
        manifest = {
            "batch_id": batch,
            "updated_at": now_iso(),
            "files": self._file_rows(list(stage.glob("*.sswweb"))),
        }
        write_json_atomic(stage / "manifest.json", manifest)
        return {
            "batch_id": batch,
            "file": manifest["files"][-1] if manifest["files"] else {},
            "staged_count": len(manifest["files"]),
        }

    def commit_base_batch(self, context: Any, batch_id: Any, expected_count: Any = None) -> dict[str, Any]:
        batch = self._validate_batch_id(batch_id)
        stage = (Path(context.state_root) / "base_staging" / batch).resolve()
        staging_root = (Path(context.state_root) / "base_staging").resolve()
        if staging_root not in stage.parents or not stage.is_dir():
            raise ValueError("O lote temporário da Base SSW não foi encontrado.")
        files = sorted(stage.glob("*.sswweb"))
        if not files:
            raise ValueError("Nenhum arquivo .sswweb foi enviado para o lote.")
        if expected_count not in (None, "") and int(expected_count) != len(files):
            raise ValueError("O número de arquivos recebidos não corresponde ao conjunto selecionado.")

        validation = context.xml_service.validate_base_source(stage)
        target = Path(context.upload_categories["bases"]).resolve()
        target.parent.mkdir(parents=True, exist_ok=True)
        history = Path(context.state_root) / "base_history" / datetime.now().strftime("%Y%m%d_%H%M%S")
        previous = target.parent / f".bases_previous_{uuid.uuid4().hex}"
        incoming = target.parent / f".bases_incoming_{uuid.uuid4().hex}"
        if incoming.exists():
            shutil.rmtree(incoming)
        shutil.copytree(stage, incoming, ignore=shutil.ignore_patterns("manifest.json"))
        try:
            if target.exists():
                target.replace(previous)
            incoming.replace(target)
            if previous.exists():
                history.parent.mkdir(parents=True, exist_ok=True)
                previous.replace(history)
        except Exception:
            if target.exists() and not any(target.iterdir()):
                shutil.rmtree(target, ignore_errors=True)
            if previous.exists() and not target.exists():
                previous.replace(target)
            raise
        finally:
            shutil.rmtree(stage, ignore_errors=True)
            shutil.rmtree(incoming, ignore_errors=True)

        context.xml_service.invalidate_dependencies()
        context.xml_service.clear_results()
        context.invoice_service.clear_results()
        result = self.base_overview(context)
        result.update({
            "validated_rows": int(validation.get("row_count") or 0),
            "validated_files": int(validation.get("file_count") or len(files)),
            "replaced_at": now_iso(),
            "previous_backup": str(history) if history.exists() else "",
            "requires_reprocessing": True,
        })
        return result

    @staticmethod
    def _xlsx_sheet_names(path: Path) -> list[str]:
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(path, "r") as archive:
            bad = archive.testzip()
            if bad:
                raise ValueError(f"A planilha XLSX está corrompida na entrada {bad}.")
            root = ET.fromstring(archive.read("xl/workbook.xml"))
        sheets = root.find(f"{{{main_ns}}}sheets")
        return [str(item.attrib.get("name") or "") for item in list(sheets) if sheets is not None] if sheets is not None else []

    @staticmethod
    def _openpyxl():
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError(
                "O módulo openpyxl é necessário para separar e compilar as tabelas de parceiros."
            ) from exc
        return load_workbook

    @staticmethod
    def _sheet_partner_column(sheet: Any) -> int | None:
        for index, cell in enumerate(sheet[1], start=1):
            if str(cell.value or "").strip().casefold() == "parceiro id":
                return index
        return None

    def _legacy_partner_source(self, context: Any) -> Path:
        uploaded_root = Path(context.upload_categories["tabelas"]).resolve()
        candidates = sorted(
            (path for path in uploaded_root.glob("*.xlsx") if path.is_file()),
            key=lambda path: path.stat().st_mtime_ns,
            reverse=True,
        )
        if candidates:
            return candidates[0]
        official = self.project_root / "tabelas" / "cadastro_tabelas_parceiros.xlsx"
        if not official.is_file():
            raise FileNotFoundError("A tabela oficial de parceiros não foi localizada.")
        return official.resolve()

    def _partner_identity(self, path: Path) -> tuple[str, str, int]:
        load_workbook = self._openpyxl()
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            if "PARCEIROS" not in workbook.sheetnames:
                raise ValueError("O arquivo do parceiro não possui a aba PARCEIROS.")
            sheet = workbook["PARCEIROS"]
            headers = {str(cell.value or "").strip(): index for index, cell in enumerate(sheet[1], start=1)}
            id_col = headers.get("Parceiro ID")
            name_col = headers.get("Nome Parceiro")
            if not id_col:
                raise ValueError("A aba PARCEIROS não possui a coluna Parceiro ID.")
            identities: list[tuple[str, str]] = []
            for row in range(2, sheet.max_row + 1):
                partner_id = str(sheet.cell(row, id_col).value or "").strip()
                if not partner_id:
                    continue
                name = str(sheet.cell(row, name_col).value or partner_id).strip() if name_col else partner_id
                identities.append((partner_id, name))
            unique = {item[0] for item in identities}
            if len(unique) != 1:
                raise ValueError("Cada arquivo deve conter exatamente um Parceiro ID na aba PARCEIROS.")
            partner_id, name = identities[0]
            for worksheet in workbook.worksheets:
                partner_col = self._sheet_partner_column(worksheet)
                if not partner_col:
                    continue
                found = {
                    str(worksheet.cell(row, partner_col).value or "").strip()
                    for row in range(2, worksheet.max_row + 1)
                    if str(worksheet.cell(row, partner_col).value or "").strip()
                }
                if found.difference({partner_id}):
                    raise ValueError(
                        f"A aba {worksheet.title} contém dados de outro parceiro: "
                        + ", ".join(sorted(found.difference({partner_id})))
                    )
            rules = 0
            if "REGRAS_PERCENTUAL" in workbook.sheetnames:
                rules_sheet = workbook["REGRAS_PERCENTUAL"]
                rules = sum(
                    1 for row in range(2, rules_sheet.max_row + 1)
                    if any(rules_sheet.cell(row, col).value not in (None, "") for col in range(1, rules_sheet.max_column + 1))
                )
            return partner_id, name, rules
        finally:
            workbook.close()

    def _split_partner_workbook(self, source: Path, destination_root: Path) -> list[Path]:
        load_workbook = self._openpyxl()
        source = Path(source).resolve()
        workbook = load_workbook(source, read_only=False, data_only=False)
        try:
            if "PARCEIROS" not in workbook.sheetnames:
                raise ValueError("A tabela oficial não possui a aba PARCEIROS.")
            sheet = workbook["PARCEIROS"]
            partner_col = self._sheet_partner_column(sheet)
            if not partner_col:
                raise ValueError("A tabela oficial não possui a coluna Parceiro ID.")
            partner_ids = []
            for row in range(2, sheet.max_row + 1):
                partner_id = str(sheet.cell(row, partner_col).value or "").strip()
                if partner_id and partner_id not in partner_ids:
                    partner_ids.append(partner_id)
        finally:
            workbook.close()
        if not partner_ids:
            raise ValueError("A tabela oficial não contém parceiros válidos.")

        destination_root.mkdir(parents=True, exist_ok=True)
        generated: list[Path] = []
        for partner_id in partner_ids:
            partner_book = load_workbook(source, read_only=False, data_only=False)
            try:
                for worksheet in partner_book.worksheets:
                    partner_index = self._sheet_partner_column(worksheet)
                    if not partner_index:
                        continue
                    for row in range(worksheet.max_row, 1, -1):
                        value = str(worksheet.cell(row, partner_index).value or "").strip()
                        if value != partner_id:
                            worksheet.delete_rows(row, 1)
                safe_id = safe_name(partner_id, "PARCEIRO").replace(" ", "_")
                target = destination_root / f"{safe_id}.xlsx"
                partner_book.save(target)
                self._partner_identity(target)
                generated.append(target)
            finally:
                partner_book.close()
        return generated

    @staticmethod
    def _copy_row(source_sheet: Any, source_row: int, target_sheet: Any, target_row: int) -> None:
        for column in range(1, source_sheet.max_column + 1):
            source = source_sheet.cell(source_row, column)
            target = target_sheet.cell(target_row, column, source.value)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.protection:
                target.protection = copy(source.protection)
        if source_row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height

    def _build_partner_aggregate(self, context: Any, files: list[Path], target: Path) -> dict[str, Any]:
        if not files:
            raise ValueError("É necessário manter pelo menos um arquivo de parceiro.")
        load_workbook = self._openpyxl()
        template = self._legacy_partner_source(context)
        workbook = load_workbook(template, read_only=False, data_only=False)
        try:
            # R12.13: a regra especial da Graúna passa a declarar R$/kg como
            # dado estruturado. Consolidados antigos não possuíam esta coluna;
            # acrescentá-la aqui mantém a compilação atômica compatível com os
            # arquivos individuais já existentes.
            if "REGRAS_PESO_ESPECIAL" in workbook.sheetnames:
                weight_sheet = workbook["REGRAS_PESO_ESPECIAL"]
                normalized_headers = {
                    re.sub(r"[^A-Z0-9]", "", str(weight_sheet.cell(1, column).value or "").upper()): column
                    for column in range(1, weight_sheet.max_column + 1)
                }
                if "VALORKG" not in normalized_headers:
                    weight_sheet.cell(1, weight_sheet.max_column + 1, "Valor KG")

            scoped_sheets: dict[str, int] = {}
            for worksheet in workbook.worksheets:
                partner_col = self._sheet_partner_column(worksheet)
                if not partner_col:
                    continue
                scoped_sheets[worksheet.title] = partner_col
                if worksheet.max_row > 1:
                    worksheet.delete_rows(2, worksheet.max_row - 1)
            for partner_path in sorted(files, key=lambda item: item.name.casefold()):
                partner_book = load_workbook(partner_path, read_only=False, data_only=False)
                try:
                    self._partner_identity(partner_path)
                    for sheet_name in scoped_sheets:
                        if sheet_name not in partner_book.sheetnames:
                            continue
                        source_sheet = partner_book[sheet_name]
                        target_sheet = workbook[sheet_name]
                        for row in range(2, source_sheet.max_row + 1):
                            if not any(source_sheet.cell(row, col).value not in (None, "") for col in range(1, source_sheet.max_column + 1)):
                                continue
                            self._copy_row(source_sheet, row, target_sheet, target_sheet.max_row + 1)
                finally:
                    partner_book.close()
            target.parent.mkdir(parents=True, exist_ok=True)
            temporary = target.with_name(f".{target.stem}_{uuid.uuid4().hex}.xlsx")
            workbook.save(temporary)
        finally:
            workbook.close()
        try:
            validation = context.xml_service.validate_table_source(temporary)
            os.replace(temporary, target)
        finally:
            temporary.unlink(missing_ok=True)
        return validation

    def ensure_partner_files(self, context: Any) -> dict[str, Any]:
        with self._partner_lock:
            files = sorted(self.partner_files_root.glob("*.xlsx"))
            if not files:
                source = self._legacy_partner_source(context)
                staging = self.partner_root / f".split_{uuid.uuid4().hex}"
                try:
                    files = self._split_partner_workbook(source, staging)
                    for generated in files:
                        os.replace(generated, self.partner_files_root / generated.name)
                finally:
                    shutil.rmtree(staging, ignore_errors=True)
                files = sorted(self.partner_files_root.glob("*.xlsx"))
            source_signature = "\n".join(
                f"{path.name}|{path.stat().st_size}|{sha256_file(path)}" for path in files
            )
            signature_path = self.partner_root / "compiled_signature.txt"
            old_signature = signature_path.read_text(encoding="utf-8") if signature_path.is_file() else ""
            if not self.partner_aggregate_path.is_file() or source_signature != old_signature:
                validation = self._build_partner_aggregate(context, files, self.partner_aggregate_path)
                signature_path.write_text(source_signature, encoding="utf-8")
                context.xml_service.invalidate_dependencies()
            else:
                validation = context.xml_service.validate_table_source(self.partner_aggregate_path)
            return {
                "file_count": len(files),
                "aggregate": str(self.partner_aggregate_path),
                "partners": int(validation.get("partner_count") or 0),
                "rules": int(validation.get("rule_count") or 0),
            }

    def partner_files_overview(self, context: Any) -> list[dict[str, Any]]:
        self.ensure_partner_files(context)
        rows: list[dict[str, Any]] = []
        for path in sorted(self.partner_files_root.glob("*.xlsx"), key=lambda item: item.name.casefold()):
            partner_id, name, rules = self._partner_identity(path)
            stat = path.stat()
            rows.append({
                "partner_id": partner_id,
                "name": name,
                "filename": path.name,
                "path": str(path.resolve()),
                "rules": rules,
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
                "sha256": sha256_file(path),
            })
        return rows

    def import_partner_file(self, context: Any, payload: bytes, filename: Any) -> dict[str, Any]:
        name = safe_name(filename, "parceiro.xlsx")
        if Path(name).suffix.lower() != ".xlsx":
            raise ValueError("O arquivo do parceiro deve ser uma planilha XLSX.")
        if not payload or len(payload) > MAX_PARTNER_TABLE_BYTES:
            raise ValueError("A tabela deve ter conteúdo e no máximo 40 MB.")
        with self._partner_lock:
            self.ensure_partner_files(context)
            staging = self.partner_root / f".partner_{uuid.uuid4().hex}.xlsx"
            staging.write_bytes(payload)
            candidate_aggregate = self.partner_root / f".aggregate_{uuid.uuid4().hex}.xlsx"
            try:
                partner_id, partner_name, rules = self._partner_identity(staging)
                target_name = safe_name(partner_id, "PARCEIRO").replace(" ", "_") + ".xlsx"
                target = self.partner_files_root / target_name
                candidates = [path for path in self.partner_files_root.glob("*.xlsx") if path.resolve() != target.resolve()]
                candidates.append(staging)
                validation = self._build_partner_aggregate(context, candidates, candidate_aggregate)
                backup = None
                if target.is_file():
                    backup = self.partner_history_root / datetime.now().strftime("%Y%m%d_%H%M%S") / target.name
                    backup.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(target, backup)
                os.replace(staging, target)
                os.replace(candidate_aggregate, self.partner_aggregate_path)
                (self.partner_root / "compiled_signature.txt").unlink(missing_ok=True)
                self.ensure_partner_files(context)
                context.xml_service.invalidate_dependencies()
                context.xml_service.clear_results()
                return {
                    "partner_id": partner_id,
                    "name": partner_name,
                    "filename": target.name,
                    "rules": rules,
                    "partners": int(validation.get("partner_count") or 0),
                    "total_rules": int(validation.get("rule_count") or 0),
                    "sha256": sha256_file(target),
                    "backup": str(backup) if backup else "",
                    "requires_reprocessing": True,
                    "updated_at": now_iso(),
                }
            finally:
                staging.unlink(missing_ok=True)
                candidate_aggregate.unlink(missing_ok=True)

    def delete_partner_file(self, context: Any, partner_id: Any) -> dict[str, Any]:
        normalized = str(partner_id or "").strip()
        if not normalized:
            raise ValueError("Parceiro não informado.")
        with self._partner_lock:
            rows = self.partner_files_overview(context)
            selected = next((row for row in rows if row["partner_id"] == normalized), None)
            if selected is None:
                raise KeyError("Arquivo do parceiro não encontrado.")
            if len(rows) <= 1:
                raise ValueError("Não é possível excluir o último parceiro ativo.")
            target = Path(selected["path"])
            candidates = [path for path in self.partner_files_root.glob("*.xlsx") if path.resolve() != target.resolve()]
            candidate_aggregate = self.partner_root / f".aggregate_{uuid.uuid4().hex}.xlsx"
            try:
                validation = self._build_partner_aggregate(context, candidates, candidate_aggregate)
                backup = self.partner_history_root / datetime.now().strftime("%Y%m%d_%H%M%S") / target.name
                backup.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target, backup)
                target.unlink()
                os.replace(candidate_aggregate, self.partner_aggregate_path)
                (self.partner_root / "compiled_signature.txt").unlink(missing_ok=True)
                self.ensure_partner_files(context)
                context.xml_service.invalidate_dependencies()
                context.xml_service.clear_results()
                return {
                    "deleted": True,
                    "partner_id": normalized,
                    "backup": str(backup),
                    "partners": int(validation.get("partner_count") or 0),
                    "rules": int(validation.get("rule_count") or 0),
                    "requires_reprocessing": True,
                }
            finally:
                candidate_aggregate.unlink(missing_ok=True)

    def export_partner_registration_template(self, context: Any) -> Path:
        """Gera um modelo sanitizado e guiado para cadastrar um parceiro.

        A planilha mantém os nomes de abas/colunas aceitos pelo motor, mas não
        copia dados comerciais reais. Os exemplos ficam em uma aba separada,
        evitando que sejam importados por engano como regras oficiais.
        """
        with self._partner_lock:
            source = self.partner_aggregate_path if self.partner_aggregate_path.is_file() else self._legacy_partner_source(context)
            load_workbook = self._openpyxl()
            workbook = load_workbook(source, read_only=False, data_only=False)
            try:
                # Remove todos os dados reais, preservando cabeçalhos, estilos e
                # a estrutura aceita pela engine RC26.6.
                for worksheet in workbook.worksheets:
                    if worksheet.max_row > 1:
                        worksheet.delete_rows(2, worksheet.max_row - 1)

                if "INSTRUCOES" in workbook.sheetnames:
                    instructions = workbook["INSTRUCOES"]
                else:
                    instructions = workbook.create_sheet("INSTRUCOES", 0)
                instructions.delete_rows(1, max(1, instructions.max_row))
                instruction_rows = [
                    ("MODELO DE CADASTRO DE PARCEIRO — CENTRAL CT-e", "Preencha um parceiro por arquivo e importe pela aba Parceiros / Tabelas."),
                    ("1. Cadastro", "Preencha a aba PARCEIROS. Parceiro ID e Nome Parceiro são obrigatórios."),
                    ("2. Regra principal", "Preencha ao menos uma linha em REGRAS_PERCENTUAL com Regra ID, Parceiro ID, destino/região e percentual ou frete mínimo."),
                    ("3. Abas opcionais", "Use REGIOES, REGRAS_EXTRAS, REGRAS_PESO_ESPECIAL e ALIAS_PARCEIROS somente quando a regra estiver prevista na engine."),
                    ("4. Status Revisão", "Use REVISAR_OK quando os dados estiverem conferidos. PENDENTE_* permanece visível no card Em revisão."),
                    ("5. Regra não suportada", "Não improvise novas fórmulas na planilha. Encaminhe a regra comercial para atualização da engine antes da importação."),
                    ("6. Importação", "Salve em XLSX. O sistema valida o arquivo inteiro e só publica após a compilação atômica."),
                    ("Exemplos", "Consulte a aba MODELO_EXEMPLOS. Ela é somente orientativa e não entra na tabela oficial."),
                ]
                for row_index, row in enumerate(instruction_rows, 1):
                    instructions.cell(row_index, 1, row[0])
                    instructions.cell(row_index, 2, row[1])
                instructions.column_dimensions["A"].width = 34
                instructions.column_dimensions["B"].width = 108
                instructions.freeze_panes = "A2"

                if "MODELO_EXEMPLOS" in workbook.sheetnames:
                    workbook.remove(workbook["MODELO_EXEMPLOS"])
                examples = workbook.create_sheet("MODELO_EXEMPLOS", 1)
                examples.append(["Aba", "Campo", "Exemplo", "Obrigatório", "Orientação"])
                example_rows = [
                    ("PARCEIROS", "Parceiro ID", "TRANSPORTADORA_X", "Sim", "Identificador único, sem alterar depois da publicação."),
                    ("PARCEIROS", "Nome Parceiro", "Transportadora Exemplo Ltda", "Sim", "Nome comercial exibido no sistema."),
                    ("PARCEIROS", "CNPJ", "00111222000133", "Recomendado", "Somente números; ajuda a localizar o parceiro no XML."),
                    ("PARCEIROS", "Nome no XML / Alias principal", "TRANSPORTADORA EXEMPLO", "Recomendado", "Nome como aparece no emitente do CT-e."),
                    ("REGRAS_PERCENTUAL", "Regra ID", "TX_PA_CAPITAL", "Sim", "Identificador único da regra dentro do parceiro."),
                    ("REGRAS_PERCENTUAL", "Parceiro ID", "TRANSPORTADORA_X", "Sim", "Deve ser exatamente igual ao Parceiro ID da aba PARCEIROS."),
                    ("REGRAS_PERCENTUAL", "Destino Cidade", "BELÉM", "Conforme regra", "Informe cidade e UF ou uma Região / Base."),
                    ("REGRAS_PERCENTUAL", "Destino UF", "PA", "Conforme regra", "Use a sigla da UF para evitar cidades homônimas."),
                    ("REGRAS_PERCENTUAL", "Percentual", "7,5%", "Percentual ou mínimo", "Aceita 7,5%, 0,075 ou 7,5 para representar 7,5%."),
                    ("REGRAS_PERCENTUAL", "Frete Mínimo", "85,00", "Percentual ou mínimo", "Valor mínimo aplicado pela regra."),
                    ("REGRAS_PERCENTUAL", "Base Cálculo", "ORIGINAL", "Sim", "Use apenas opções já reconhecidas pelo motor."),
                    ("REGRAS_PERCENTUAL", "Status Revisão", "REVISAR_OK", "Sim", "Somente marque OK após conferir a proposta/tabela comercial."),
                    ("REGIOES", "Região/Base", "PA_CAPITAL", "Quando aplicável", "Agrupa cidades que compartilham a mesma regra."),
                    ("REGIOES", "Cidade", "ANANINDEUA", "Quando aplicável", "Uma cidade por linha, sempre com UF."),
                    ("REGRAS_EXTRAS", "Tipo Extra", "GRIS", "Quando suportado", "Use somente tipos já implementados na engine."),
                    ("ALIAS_PARCEIROS", "Nome no XML", "TRANSPORTADORA X", "Opcional", "Cadastre variações reais encontradas no XML."),
                ]
                for row in example_rows:
                    examples.append(list(row))
                examples.freeze_panes = "A2"
                for column, width in {"A": 24, "B": 32, "C": 30, "D": 18, "E": 72}.items():
                    examples.column_dimensions[column].width = width

                # Destaques e instruções diretamente nos cabeçalhos das abas de
                # preenchimento. A formatação é deliberadamente simples.
                try:
                    from openpyxl.comments import Comment
                    from openpyxl.styles import Font, PatternFill, Alignment
                except Exception:
                    Comment = None
                    Font = PatternFill = Alignment = None

                header_fill = PatternFill("solid", fgColor="1F4E78") if PatternFill else None
                required_fill = PatternFill("solid", fgColor="FFF2CC") if PatternFill else None
                header_font = Font(color="FFFFFF", bold=True) if Font else None
                title_font = Font(bold=True, size=14) if Font else None
                required_by_sheet = {
                    "PARCEIROS": {"Parceiro ID", "Nome Parceiro"},
                    "REGRAS_PERCENTUAL": {"Regra ID", "Parceiro ID", "Base Cálculo", "Status Revisão"},
                    "REGIOES": {"Região ID", "Parceiro ID", "Região/Base", "Cidade", "UF", "Status Revisão"},
                    "REGRAS_EXTRAS": {"Extra ID", "Parceiro ID", "Tipo Extra", "Status Revisão"},
                    "REGRAS_PESO_ESPECIAL": {"Regra ID", "Parceiro ID", "Peso Mínimo KG", "Modo Cálculo"},
                    "ALIAS_PARCEIROS": {"Parceiro ID", "Nome no XML"},
                }
                guidance = {
                    "Parceiro ID": "Repita exatamente o mesmo identificador em todas as abas.",
                    "Regra ID": "Identificador único da regra.",
                    "Percentual": "Informe percentual comercial; use o padrão apresentado na aba MODELO_EXEMPLOS.",
                    "Frete Mínimo": "Preencha quando houver valor mínimo na proposta.",
                    "Base Cálculo": "Não crie opções novas sem atualização da engine.",
                    "Status Revisão": "REVISAR_OK = conferido; PENDENTE_* = ainda exige revisão.",
                }
                for worksheet in workbook.worksheets:
                    if worksheet.title in {"INSTRUCOES", "MODELO_EXEMPLOS"}:
                        continue
                    worksheet.freeze_panes = "A2"
                    for cell in worksheet[1]:
                        if not str(cell.value or "").strip():
                            continue
                        if header_fill:
                            cell.fill = header_fill
                        if header_font:
                            cell.font = header_font
                        if Alignment:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        if Comment and str(cell.value) in guidance:
                            cell.comment = Comment(guidance[str(cell.value)], "Central CT-e")
                        worksheet.column_dimensions[cell.column_letter].width = max(14, min(30, len(str(cell.value)) + 4))
                    required = required_by_sheet.get(worksheet.title, set())
                    for cell in worksheet[1]:
                        if str(cell.value or "") in required and required_fill:
                            cell.fill = required_fill
                            if Font:
                                cell.font = Font(color="7F6000", bold=True)

                if title_font:
                    instructions["A1"].font = title_font
                workbook.active = workbook.sheetnames.index("INSTRUCOES")

                target = self.partner_export_root / "MODELO_CADASTRO_PARCEIRO_CENTRAL_CTE.xlsx"
                target.parent.mkdir(parents=True, exist_ok=True)
                temporary = target.with_name(f".{target.stem}_{uuid.uuid4().hex}.xlsx")
                workbook.save(temporary)
                os.replace(temporary, target)
                return target
            finally:
                workbook.close()

    def export_partner_files_zip(self, context: Any) -> Path:
        rows = self.partner_files_overview(context)
        target = self.partner_export_root / f"tabelas_parceiros_separadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        target.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
            for row in rows:
                path = Path(row["path"])
                archive.write(path, f"parceiros/{path.name}")
            archive.writestr(
                "LEIA_ME.txt",
                "Cada arquivo XLSX representa exatamente um parceiro. Edite ou substitua somente o arquivo correspondente.\n",
            )
        return target

    def partner_file(self, context: Any, partner_id: Any) -> Path:
        normalized = str(partner_id or "").strip()
        row = next((item for item in self.partner_files_overview(context) if item["partner_id"] == normalized), None)
        if row is None:
            raise FileNotFoundError("Arquivo do parceiro não encontrado.")
        return Path(row["path"]).resolve()

    def replace_partner_table(self, context: Any, payload: bytes, filename: Any) -> dict[str, Any]:
        """Compatibilidade: importa uma planilha completa e a divide por parceiro."""
        name = safe_name(filename, "cadastro_tabelas_parceiros.xlsx")
        if Path(name).suffix.lower() != ".xlsx":
            raise ValueError("A tabela de parceiros deve ser um arquivo XLSX.")
        if not payload or len(payload) > MAX_PARTNER_TABLE_BYTES:
            raise ValueError("A tabela deve ter conteúdo e no máximo 40 MB.")
        with self._partner_lock:
            temporary = self.partner_root / f".monolithic_{uuid.uuid4().hex}.xlsx"
            staging = self.partner_root / f".split_{uuid.uuid4().hex}"
            temporary.write_bytes(payload)
            try:
                validation = context.xml_service.validate_table_source(temporary)
                compatibility_root = Path(context.upload_categories["tabelas"]).resolve()
                compatibility_root.mkdir(parents=True, exist_ok=True)
                compatibility_file = compatibility_root / "cadastro_tabelas_parceiros.xlsx"
                compatibility_file.write_bytes(payload)
                generated = self._split_partner_workbook(temporary, staging)
                if self.partner_files_root.exists() and any(self.partner_files_root.glob("*.xlsx")):
                    backup_root = self.partner_history_root / datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_root.mkdir(parents=True, exist_ok=True)
                    for old in self.partner_files_root.glob("*.xlsx"):
                        shutil.copy2(old, backup_root / old.name)
                        old.unlink()
                for item in generated:
                    os.replace(item, self.partner_files_root / item.name)
                self._build_partner_aggregate(context, sorted(self.partner_files_root.glob("*.xlsx")), self.partner_aggregate_path)
                (self.partner_root / "compiled_signature.txt").unlink(missing_ok=True)
                self.ensure_partner_files(context)
                context.xml_service.invalidate_dependencies()
                context.xml_service.clear_results()
                return {
                    "path": str(self.partner_aggregate_path),
                    "filename": self.partner_aggregate_path.name,
                    "sha256": sha256_file(self.partner_aggregate_path),
                    "partners": int(validation.get("partner_count") or 0),
                    "rules": int(validation.get("rule_count") or 0),
                    "separate_files": len(list(self.partner_files_root.glob("*.xlsx"))),
                    "requires_reprocessing": True,
                    "updated_at": now_iso(),
                }
            finally:
                temporary.unlink(missing_ok=True)
                shutil.rmtree(staging, ignore_errors=True)

    def active_partner_table(self, context: Any) -> Path:
        try:
            published = context.xml_service.resolve_table_source(raise_on_missing=False)
            if published and Path(published).is_file():
                published_path = Path(published).resolve()
                if published_path != self.partner_aggregate_path.resolve():
                    return published_path
        except Exception:
            pass
        self.ensure_partner_files(context)
        if not self.partner_aggregate_path.is_file():
            raise FileNotFoundError("A tabela compilada de parceiros não foi localizada.")
        return self.partner_aggregate_path.resolve()

    def export_qa_bundle(self, context: Any) -> Path:
        """Exporta o caderno global junto com as imagens anexadas.

        O JSON isolado mantém apenas metadados e URLs internas. O ZIP inclui
        os bytes reais em ``attachments/`` e adiciona ``export_path`` a cada
        anexo, permitindo leitura fora do sistema e por ferramentas de QA.
        """
        notes = read_json(Path(context.qa_path), [])
        if not isinstance(notes, list):
            notes = []
        exported_notes: list[dict[str, Any]] = []
        attachment_files: list[tuple[Path, str]] = []
        manifest_files: list[dict[str, Any]] = []
        for item in notes:
            if not isinstance(item, Mapping):
                continue
            cloned = json.loads(json.dumps(dict(item), ensure_ascii=False, default=str))
            attachment = cloned.get("attachment")
            if isinstance(attachment, dict):
                attachment_id = re.sub(r"[^A-Za-z0-9_-]+", "", str(attachment.get("id") or ""))[:80]
                matches = sorted(self.qa_attachment_root.glob(f"{attachment_id}.*")) if attachment_id else []
                source = matches[0] if len(matches) == 1 and matches[0].is_file() else None
                if source is not None:
                    export_name = f"attachments/{source.name}"
                    attachment["export_path"] = export_name
                    attachment["included_in_export"] = True
                    attachment_files.append((source, export_name))
                    manifest_files.append({
                        "path": export_name,
                        "size_bytes": source.stat().st_size,
                        "sha256": sha256_file(source),
                        "note_id": cloned.get("id"),
                    })
                else:
                    attachment["included_in_export"] = False
                    attachment["export_error"] = "Arquivo de imagem não localizado no servidor."
            exported_notes.append(cloned)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = self.qa_export_root / f"central_cte_qa_com_imagens_{timestamp}.zip"
        payload = json.dumps(exported_notes, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        manifest = {
            "generated_at": now_iso(),
            "notes": len(exported_notes),
            "attachments_included": len(attachment_files),
            "format": "Central CT-e QA bundle v1",
            "files": [
                {"path": "central_cte_qa.json", "size_bytes": len(payload), "sha256": sha256_bytes(payload)},
                *manifest_files,
            ],
        }
        readme = (
            "CENTRAL CT-e / DACTE - Caderno de homologacao com imagens\n\n"
            "Abra central_cte_qa.json. Cada attachment.export_path aponta para a imagem\n"
            "correspondente dentro da pasta attachments/ deste ZIP.\n"
        ).encode("utf-8")
        with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
            archive.writestr("central_cte_qa.json", payload)
            archive.writestr("manifest_sha256.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"))
            archive.writestr("LEIA-ME.txt", readme)
            seen: set[str] = set()
            for source, export_name in attachment_files:
                if export_name in seen:
                    continue
                seen.add(export_name)
                archive.write(source, export_name)
        return target.resolve()

    def clear_qa(self, context: Any) -> dict[str, Any]:
        notes_path = Path(context.qa_path)
        notes = read_json(notes_path, [])
        if not isinstance(notes, list):
            notes = []
        archive = notes_path.parent / "archive" / f"qa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        if notes:
            archive.parent.mkdir(parents=True, exist_ok=True)
            write_json_atomic(archive, notes)
        deleted_ids = [str(item.get("id")) for item in notes if isinstance(item, Mapping) and item.get("id")]
        write_json_atomic(notes_path, [])

        # Os cadernos antigos por workspace eram a fonte de ressurreição de
        # ocorrências já apagadas após reiniciar o servidor. Esvaziá-los mantém
        # o arquivo de migração como evidência, mas impede nova importação.
        legacy_cleared = 0
        for legacy_path in self.workspaces_root.glob("*/state/qa_notes.json"):
            legacy_rows = read_json(legacy_path, [])
            if isinstance(legacy_rows, list) and legacy_rows:
                migrated = legacy_path.with_name("qa_notes_migrated_mvp13.json")
                if not migrated.exists():
                    shutil.copy2(legacy_path, migrated)
                write_json_atomic(legacy_path, [])
                legacy_cleared += len(legacy_rows)

        write_json_atomic(
            notes_path.parent / "last_clear.json",
            {"cleared_at": now_iso(), "deleted_ids": deleted_ids, "legacy_rows_cleared": legacy_cleared},
        )
        return {
            "deleted": len(notes),
            "legacy_rows_cleared": legacy_cleared,
            "archived": bool(notes),
            "archive_path": str(archive) if notes else "",
            "cleared_at": now_iso(),
        }
